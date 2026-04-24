#!/usr/bin/env python3
"""Batch monthly return extractor for Caissa fund return endpoints.

This script reads a list of funds from CSV, TSV, or XLSX input, resolves each
fund against the v0 fund catalog, selects a return source, and exports monthly
return rows to CSV.

Authentication is not handled here directly. Pass an existing bearer token via
--token or the CAISSA_BEARER_TOKEN environment variable.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - depends on local runtime
    load_workbook = None


DEFAULT_BASE_URL = "https://client-api.caissallc.com"
DEFAULT_PAGE_SIZE = 500

INPUT_HEADER_ALIASES = {
    "fundcode": "fundCode",
    "code": "fundCode",
    "fundname": "fundName",
    "name": "fundName",
    "returnsourcetype": "returnSourceType",
    "sourcetype": "returnSourceType",
    "returnsourcenamecontains": "returnSourceNameContains",
    "returnsourcename": "returnSourceNameContains",
    "sourcenamecontains": "returnSourceNameContains",
    "returnsourceid": "returnSourceId",
    "sourceid": "returnSourceId",
    "notes": "notes",
}

RETURN_OUTPUT_HEADERS = [
    "inputRowNumber",
    "inputFundCode",
    "inputFundName",
    "matchedFundId",
    "matchedFundCode",
    "matchedFundName",
    "entityId",
    "entityType",
    "returnSourceId",
    "returnSourceType",
    "returnSourceName",
    "returnSourceSubTitle",
    "date",
    "returnRate",
    "notes",
]

ISSUE_OUTPUT_HEADERS = [
    "inputRowNumber",
    "inputFundCode",
    "inputFundName",
    "issueType",
    "details",
    "matchedCandidates",
    "notes",
]


class BatchError(Exception):
    """Base exception for predictable batch failures."""


class InputValidationError(BatchError):
    """Raised when the input file or CLI arguments are invalid."""


class ApiRequestError(BatchError):
    """Raised when an API request fails."""


class AmbiguousMatchError(BatchError):
    """Raised when a row matches more than one fund or return source."""


@dataclass
class FundInputRow:
    row_number: int
    fund_code: str
    fund_name: str
    return_source_type: str
    return_source_name_contains: str
    return_source_id: str
    notes: str


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read a CSV/TSV/XLSX list of funds and export monthly return rows "
            "from the Caissa v0 fund returns endpoint."
        )
    )
    parser.add_argument("--input", required=True, help="Input CSV, TSV, or XLSX path.")
    parser.add_argument(
        "--output",
        required=True,
        help="Output CSV path for monthly return rows.",
    )
    parser.add_argument(
        "--issues-output",
        help="Optional CSV path for unresolved rows and warnings.",
    )
    parser.add_argument(
        "--date-from",
        required=True,
        help="Inclusive start timestamp, for example 2022-01-01T00:00:00Z.",
    )
    parser.add_argument(
        "--date-to",
        required=True,
        help="Inclusive end timestamp, for example 2026-03-31T23:59:59Z.",
    )
    parser.add_argument(
        "--token",
        help="Bearer token. If omitted, the script uses CAISSA_BEARER_TOKEN.",
    )
    parser.add_argument(
        "--sheet",
        help="Worksheet name for XLSX input. Defaults to the first sheet.",
    )
    parser.add_argument(
        "--base-url",
        default=DEFAULT_BASE_URL,
        help=f"API base URL. Defaults to {DEFAULT_BASE_URL}.",
    )
    parser.add_argument(
        "--page-size",
        type=int,
        default=DEFAULT_PAGE_SIZE,
        help=f"Page size for paged endpoints. Defaults to {DEFAULT_PAGE_SIZE}.",
    )
    parser.add_argument(
        "--preferred-source-types",
        default="",
        help=(
            "Optional comma-separated returnSourceType preference order, for example "
            "FundAccounting,FundHybrid. By default the script stays strict and does "
            "not auto-pick among multiple sources."
        ),
    )
    parser.add_argument(
        "--computation-methodology",
        help=(
            "Optional return computation methodology, for example TimeWeighted or "
            "ModifiedDietzMonthly."
        ),
    )
    parser.add_argument(
        "--book-closing-calendar-id",
        type=int,
        help="Optional book closing calendar id to pass through to the returns endpoint.",
    )
    parser.add_argument(
        "--timeout-seconds",
        type=int,
        default=60,
        help="Per-request timeout in seconds. Defaults to 60.",
    )
    parser.add_argument(
        "--validate-input-only",
        action="store_true",
        help="Validate the input file structure without calling the API.",
    )
    return parser.parse_args(argv)


def normalize_header(value: str) -> str:
    return "".join(ch.lower() for ch in value.strip() if ch.isalnum())


def normalize_text(value: str) -> str:
    return " ".join(value.split()).casefold()


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def split_csv_list(value: str) -> List[str]:
    return [item.strip() for item in value.split(",") if item.strip()]


def canonicalize_headers(headers: Iterable[Any]) -> Dict[int, str]:
    header_map: Dict[int, str] = {}
    seen_fields: Dict[str, int] = {}

    for index, raw_header in enumerate(headers):
        header_text = clean_cell(raw_header)
        canonical = INPUT_HEADER_ALIASES.get(normalize_header(header_text))
        if not canonical:
            continue
        if canonical in seen_fields:
            first_col = seen_fields[canonical] + 1
            raise InputValidationError(
                f"Duplicate input header mapped to '{canonical}' in columns "
                f"{first_col} and {index + 1}."
            )
        seen_fields[canonical] = index
        header_map[index] = canonical

    if "fundCode" not in header_map.values() and "fundName" not in header_map.values():
        raise InputValidationError(
            "Input file must include at least one supported identifier column: "
            "fundCode or fundName."
        )

    return header_map


def materialize_input_row(row_number: int, values: Dict[str, str]) -> Optional[FundInputRow]:
    row = FundInputRow(
        row_number=row_number,
        fund_code=values.get("fundCode", ""),
        fund_name=values.get("fundName", ""),
        return_source_type=values.get("returnSourceType", ""),
        return_source_name_contains=values.get("returnSourceNameContains", ""),
        return_source_id=values.get("returnSourceId", ""),
        notes=values.get("notes", ""),
    )

    if not any(
        [
            row.fund_code,
            row.fund_name,
            row.return_source_type,
            row.return_source_name_contains,
            row.return_source_id,
            row.notes,
        ]
    ):
        return None

    if not row.fund_code and not row.fund_name:
        raise InputValidationError(
            f"Input row {row_number} must include fundCode or fundName."
        )

    return row


def read_delimited_input(input_path: Path) -> List[FundInputRow]:
    delimiter = "\t" if input_path.suffix.lower() == ".tsv" else ","
    rows: List[FundInputRow] = []

    with input_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        try:
            raw_headers = next(reader)
        except StopIteration as exc:
            raise InputValidationError(f"Input file {input_path} is empty.") from exc

        header_map = canonicalize_headers(raw_headers)

        for row_number, raw_values in enumerate(reader, start=2):
            padded_values = list(raw_values) + [""] * max(0, len(raw_headers) - len(raw_values))
            values_by_field = {
                field_name: clean_cell(padded_values[index])
                for index, field_name in header_map.items()
            }
            row = materialize_input_row(row_number, values_by_field)
            if row:
                rows.append(row)

    return rows


def read_xlsx_input(input_path: Path, sheet_name: Optional[str]) -> List[FundInputRow]:
    if load_workbook is None:
        raise InputValidationError(
            "XLSX input requires openpyxl. Use CSV/TSV input or install openpyxl."
        )

    workbook = load_workbook(filename=input_path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise InputValidationError(
                f"Worksheet '{sheet_name}' was not found. Available sheets: "
                f"{', '.join(workbook.sheetnames)}"
            )
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration as exc:
        raise InputValidationError(f"Worksheet '{worksheet.title}' is empty.") from exc

    header_map = canonicalize_headers(raw_headers)
    rows: List[FundInputRow] = []

    for row_number, raw_values in enumerate(rows_iter, start=2):
        values_by_field = {
            field_name: clean_cell(raw_values[index] if index < len(raw_values) else "")
            for index, field_name in header_map.items()
        }
        row = materialize_input_row(row_number, values_by_field)
        if row:
            rows.append(row)

    workbook.close()
    return rows


def read_input_rows(input_path: Path, sheet_name: Optional[str]) -> List[FundInputRow]:
    if not input_path.exists():
        raise InputValidationError(f"Input file not found: {input_path}")

    suffix = input_path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        rows = read_delimited_input(input_path)
    elif suffix == ".xlsx":
        rows = read_xlsx_input(input_path, sheet_name)
    else:
        raise InputValidationError(
            "Unsupported input format. Use .csv, .tsv, or .xlsx."
        )

    if not rows:
        raise InputValidationError("No non-empty data rows were found in the input file.")

    return rows


class CaissaClient:
    def __init__(self, base_url: str, bearer_token: str, timeout_seconds: int) -> None:
        self.base_url = base_url.rstrip("/")
        self.bearer_token = bearer_token.strip()
        self.timeout_seconds = timeout_seconds

    def get_json(self, path: str, params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        query = urlencode(params or {}, doseq=True)
        url = f"{self.base_url}{path}"
        if query:
            url = f"{url}?{query}"

        request = Request(
            url,
            headers={
                "Accept": "application/json",
                "Authorization": f"Bearer {self.bearer_token}",
            },
            method="GET",
        )

        try:
            with urlopen(request, timeout=self.timeout_seconds) as response:
                payload = response.read().decode("utf-8")
        except HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            raise ApiRequestError(
                f"HTTP {exc.code} for {url}: {body[:500]}"
            ) from exc
        except URLError as exc:
            raise ApiRequestError(f"Network error for {url}: {exc.reason}") from exc

        try:
            return json.loads(payload)
        except json.JSONDecodeError as exc:
            raise ApiRequestError(
                f"Failed to parse JSON from {url}: {payload[:500]}"
            ) from exc

    def get_all_results(
        self, path: str, params: Optional[Dict[str, Any]] = None
    ) -> List[Dict[str, Any]]:
        params = dict(params or {})
        page_size = int(params.pop("PageSize", DEFAULT_PAGE_SIZE))
        page_index = 1
        results: List[Dict[str, Any]] = []

        while True:
            page_params = dict(params)
            page_params["PageIndex"] = page_index
            page_params["PageSize"] = page_size
            payload = self.get_json(path, page_params)
            page_results = payload.get("results") or []
            paging = payload.get("paging") or {}

            results.extend(page_results)

            results_size = paging.get("resultsSize")
            total_size = paging.get("totalSize")
            if total_size is not None and len(results) >= int(total_size):
                break
            if results_size is not None and int(results_size) < page_size:
                break
            if len(page_results) < page_size:
                break
            if not page_results:
                break

            page_index += 1

        return results


def build_fund_indexes(
    funds: List[Dict[str, Any]]
) -> Tuple[Dict[str, List[Dict[str, Any]]], Dict[str, List[Dict[str, Any]]]]:
    by_code: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    by_name: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    for fund in funds:
        fund_code = clean_cell(fund.get("fundCode"))
        fund_name = clean_cell(fund.get("fundName"))
        if fund_code:
            by_code[fund_code.upper()].append(fund)
        if fund_name:
            by_name[normalize_text(fund_name)].append(fund)

    return by_code, by_name


def summarize_fund_candidates(candidates: List[Dict[str, Any]]) -> str:
    return "; ".join(
        f"fundId={fund.get('fundId')} fundCode={clean_cell(fund.get('fundCode'))} "
        f"fundName={clean_cell(fund.get('fundName'))}"
        for fund in candidates
    )


def resolve_fund(
    row: FundInputRow,
    funds_by_code: Dict[str, List[Dict[str, Any]]],
    funds_by_name: Dict[str, List[Dict[str, Any]]],
) -> Dict[str, Any]:
    code_matches: List[Dict[str, Any]] = []
    name_matches: List[Dict[str, Any]] = []

    if row.fund_code:
        code_matches = funds_by_code.get(row.fund_code.upper(), [])
        if not code_matches:
            raise InputValidationError(
                f"Input row {row.row_number}: fundCode '{row.fund_code}' was not found."
            )
        if len(code_matches) > 1:
            raise AmbiguousMatchError(
                f"Input row {row.row_number}: fundCode '{row.fund_code}' matched more "
                f"than one fund: {summarize_fund_candidates(code_matches)}"
            )

    if row.fund_name:
        name_matches = funds_by_name.get(normalize_text(row.fund_name), [])
        if not name_matches:
            raise InputValidationError(
                f"Input row {row.row_number}: fundName '{row.fund_name}' was not found."
            )
        if len(name_matches) > 1:
            raise AmbiguousMatchError(
                f"Input row {row.row_number}: fundName '{row.fund_name}' matched more "
                f"than one fund. Add fundCode to disambiguate. Candidates: "
                f"{summarize_fund_candidates(name_matches)}"
            )

    if code_matches and name_matches:
        if code_matches[0].get("fundId") != name_matches[0].get("fundId"):
            raise InputValidationError(
                f"Input row {row.row_number}: fundCode '{row.fund_code}' and fundName "
                f"'{row.fund_name}' resolve to different funds."
            )
        return code_matches[0]

    if code_matches:
        return code_matches[0]

    return name_matches[0]


def summarize_return_sources(candidates: List[Dict[str, Any]]) -> str:
    return "; ".join(
        " ".join(
            [
                f"id={source.get('id')}",
                f"type={clean_cell(source.get('type'))}",
                f"returnSourceId={clean_cell(source.get('returnSourceId'))}",
                f"returnSourceType={clean_cell(source.get('returnSourceType'))}",
                f"name={clean_cell(source.get('returnSourceName') or source.get('name'))}",
            ]
        )
        for source in candidates
    )


def source_matches_name(source: Dict[str, Any], needle: str) -> bool:
    haystack = " ".join(
        [
            clean_cell(source.get("name")),
            clean_cell(source.get("returnSourceName")),
            clean_cell(source.get("returnSourceSubTitle")),
            clean_cell(source.get("returnSourceGroupName")),
        ]
    )
    return needle in normalize_text(haystack)


def select_return_source(
    row: FundInputRow,
    sources: List[Dict[str, Any]],
    preferred_source_types: List[str],
) -> Dict[str, Any]:
    if not sources:
        raise InputValidationError(
            f"Input row {row.row_number}: no return sources were found for the matched fund."
        )

    candidates = list(sources)
    filters_used: List[str] = []

    if row.return_source_id:
        target_id = row.return_source_id
        candidates = [
            source
            for source in candidates
            if clean_cell(source.get("returnSourceId")) == target_id
            or clean_cell(source.get("id")) == target_id
        ]
        filters_used.append(f"returnSourceId={target_id}")

    if row.return_source_type:
        target_type = normalize_text(row.return_source_type)
        candidates = [
            source
            for source in candidates
            if normalize_text(clean_cell(source.get("returnSourceType"))) == target_type
        ]
        filters_used.append(f"returnSourceType={row.return_source_type}")

    if row.return_source_name_contains:
        needle = normalize_text(row.return_source_name_contains)
        candidates = [
            source for source in candidates if source_matches_name(source, needle)
        ]
        filters_used.append(
            f"returnSourceNameContains={row.return_source_name_contains}"
        )

    if not candidates:
        raise InputValidationError(
            f"Input row {row.row_number}: no return source matched "
            f"{', '.join(filters_used) or 'the available sources'}."
        )

    if len(candidates) == 1:
        return candidates[0]

    if not row.return_source_type and preferred_source_types:
        for preferred_type in preferred_source_types:
            preferred_matches = [
                source
                for source in candidates
                if normalize_text(clean_cell(source.get("returnSourceType")))
                == normalize_text(preferred_type)
            ]
            if len(preferred_matches) == 1:
                return preferred_matches[0]
            if len(preferred_matches) > 1:
                raise AmbiguousMatchError(
                    f"Input row {row.row_number}: multiple return sources still match "
                    f"preferred type '{preferred_type}'. Add returnSourceId, "
                    f"returnSourceNameContains, or a more specific returnSourceType. "
                    f"Candidates: {summarize_return_sources(preferred_matches)}"
                )

    raise AmbiguousMatchError(
        f"Input row {row.row_number}: multiple return sources matched. Add "
        f"returnSourceType, returnSourceId, or returnSourceNameContains. "
        f"Candidates: {summarize_return_sources(candidates)}"
    )


def write_csv(path: Path, headers: List[str], rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def issue_row(
    row: FundInputRow, issue_type: str, details: str, matched_candidates: str = ""
) -> Dict[str, str]:
    return {
        "inputRowNumber": row.row_number,
        "inputFundCode": row.fund_code,
        "inputFundName": row.fund_name,
        "issueType": issue_type,
        "details": details,
        "matchedCandidates": matched_candidates,
        "notes": row.notes,
    }


def fetch_monthly_returns(
    client: CaissaClient,
    source: Dict[str, Any],
    args: argparse.Namespace,
) -> List[Dict[str, Any]]:
    params: Dict[str, Any] = {
        "returnSource.id": source.get("id"),
        "returnSource.type": source.get("type"),
        "returnSource.returnSourceId": source.get("returnSourceId"),
        "returnSource.returnSourceType": source.get("returnSourceType"),
        "DateFrom": args.date_from,
        "DateTo": args.date_to,
        "Periodicity": "Monthly",
        "SortBy": "Date",
        "SortOrder": "Asc",
        "PageSize": args.page_size,
    }
    if args.computation_methodology:
        params["ComputationalMethodology"] = args.computation_methodology
    if args.book_closing_calendar_id is not None:
        params["BookClosingCalendarId"] = args.book_closing_calendar_id

    return client.get_all_results("/v0/funds/returns", params=params)


def run(args: argparse.Namespace) -> int:
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    issues_output_path = (
        Path(args.issues_output).expanduser().resolve()
        if args.issues_output
        else output_path.with_name(f"{output_path.stem}_issues{output_path.suffix}")
    )

    input_rows = read_input_rows(input_path, args.sheet)
    preferred_source_types = split_csv_list(args.preferred_source_types)

    if args.validate_input_only:
        print(
            f"Validated {len(input_rows)} input rows from {input_path} "
            f"using sheet '{args.sheet or 'first sheet'}'."
        )
        return 0

    token = (args.token or os.environ.get("CAISSA_BEARER_TOKEN", "")).strip()
    if not token:
        raise InputValidationError(
            "Missing bearer token. Provide --token or set CAISSA_BEARER_TOKEN."
        )

    client = CaissaClient(
        base_url=args.base_url,
        bearer_token=token,
        timeout_seconds=args.timeout_seconds,
    )

    funds = client.get_all_results(
        "/v0/funds",
        params={"SortBy": "FundName", "SortOrder": "Asc", "PageSize": args.page_size},
    )
    funds_by_code, funds_by_name = build_fund_indexes(funds)

    source_cache: Dict[int, List[Dict[str, Any]]] = {}
    return_rows: List[Dict[str, Any]] = []
    issue_rows_out: List[Dict[str, Any]] = []

    for row in input_rows:
        try:
            fund = resolve_fund(row, funds_by_code, funds_by_name)
            fund_id = int(fund["fundId"])

            if fund_id not in source_cache:
                source_cache[fund_id] = client.get_all_results(
                    "/v0/funds/return-sources",
                    params={
                        "FundIds": [fund_id],
                        "SortBy": "Name",
                        "SortOrder": "Asc",
                        "PageSize": args.page_size,
                    },
                )

            selected_source = select_return_source(
                row=row,
                sources=source_cache[fund_id],
                preferred_source_types=preferred_source_types,
            )
            monthly_returns = fetch_monthly_returns(client, selected_source, args)

            if not monthly_returns:
                issue_rows_out.append(
                    issue_row(
                        row=row,
                        issue_type="NO_RETURNS_FOUND",
                        details=(
                            "The selected return source did not return any monthly "
                            "rows for the requested date range."
                        ),
                        matched_candidates=summarize_return_sources([selected_source]),
                    )
                )
                continue

            for item in monthly_returns:
                return_rows.append(
                    {
                        "inputRowNumber": row.row_number,
                        "inputFundCode": row.fund_code,
                        "inputFundName": row.fund_name,
                        "matchedFundId": fund.get("fundId"),
                        "matchedFundCode": clean_cell(fund.get("fundCode")),
                        "matchedFundName": clean_cell(fund.get("fundName")),
                        "entityId": clean_cell(selected_source.get("id")),
                        "entityType": clean_cell(selected_source.get("type")),
                        "returnSourceId": clean_cell(selected_source.get("returnSourceId")),
                        "returnSourceType": clean_cell(
                            selected_source.get("returnSourceType")
                        ),
                        "returnSourceName": clean_cell(
                            selected_source.get("returnSourceName")
                            or selected_source.get("name")
                        ),
                        "returnSourceSubTitle": clean_cell(
                            selected_source.get("returnSourceSubTitle")
                        ),
                        "date": clean_cell(item.get("date")),
                        "returnRate": item.get("returnRate"),
                        "notes": row.notes,
                    }
                )
        except AmbiguousMatchError as exc:
            issue_rows_out.append(
                issue_row(
                    row=row,
                    issue_type="AMBIGUOUS_MATCH",
                    details=str(exc),
                )
            )
        except InputValidationError as exc:
            issue_rows_out.append(
                issue_row(
                    row=row,
                    issue_type="INPUT_ERROR",
                    details=str(exc),
                )
            )
        except ApiRequestError as exc:
            issue_rows_out.append(
                issue_row(
                    row=row,
                    issue_type="API_ERROR",
                    details=str(exc),
                )
            )

    write_csv(output_path, RETURN_OUTPUT_HEADERS, return_rows)
    write_csv(issues_output_path, ISSUE_OUTPUT_HEADERS, issue_rows_out)

    print(f"Wrote {len(return_rows)} monthly return rows to {output_path}")
    print(f"Wrote {len(issue_rows_out)} issue rows to {issues_output_path}")

    if return_rows:
        if issue_rows_out:
            print(
                "Completed with warnings. Review the issues file for ambiguous funds, "
                "missing return sources, or empty date ranges.",
                file=sys.stderr,
            )
        return 0

    return 2


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    try:
        return run(args)
    except BatchError as exc:
        print(str(exc), file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
